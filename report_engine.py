"""
report_engine.py
=================
Bo may sinh "Bao cao ngay / Daily Report" song ngu tu file mau (template)
that, dua tren du lieu JSON do APP nhap lieu gui xuong.

Nam vung 6 tieu chi cua yeu cau:

 1. Tieng Anh truoc (dung) + Tieng Viet sau (nghieng)  -> bilingual_richtext()
 2. Du lieu nhap 1 ngon ngu van duoc dich sang song ngu -> translate() + glossary
    hoc tu chinh file mau (build_glossary) + hook goi API dich that (xem
    ham `external_translate`) khi trien khai san xuat.
 3. O to mau / chu do trong template la o nhap lieu -> khi xuat: nen trang,
    chu den -> find_input_cells() + reset_input_style()
 4. O khong nhap thi de trong, KHONG lay mac dinh cua template -> clear_or_fill()
 5. Trang anh tu dong sinh them khoi (1 dong mo ta + 2 khung anh) cho moi
    2 anh, tu dong ngat trang -> PhotoPageBuilder
 6. Tu dong wrap + chinh chieu cao dong mo ta de khong de chu chen dep
    trang in -> autofit_caption_row()

Yeu cau: openpyxl >= 3.1 (rich text), PIL (Pillow) de do kich thuoc anh.
"""

import re
import os
import copy
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter, column_index_from_string

# --------------------------------------------------------------------------
# 1) SONG NGU: TIENG ANH DUNG (truoc) + TIENG VIET NGHIENG (sau)
# --------------------------------------------------------------------------

FONT_NAME = "Arial"   # font mac dinh CHI dung khi khong doc duoc font goc cua o


_XML_INVALID_RE = re.compile(
    "[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ufdd0-\ufdef\ufffe\uffff]"
)


def _sanitize_xml_text(s: str) -> str:
    """Loai bo ky tu dieu khien KHONG hop le trong XML 1.0 — neu lot vao
    noi dung o (vd tu ket qua API dich bi loi/tra ve du lieu la) se lam
    HONG FILE .xlsx (Excel bao 'found a problem with content', da gap
    thuc te). Day la lop bao ve cuoi cung, khong phu thuoc nguon du lieu
    dau vao co sach hay khong."""
    if not s:
        return s
    return _XML_INVALID_RE.sub("", s)


def bilingual_richtext(en: str, vi: str, size: float = 11, bold: bool = False,
                        font_name: str = None):
    """Tra ve CellRichText: dong 1 = tieng Anh (dung) + dau '/' cuoi dong
    (dung quy uoc cua chinh template goc, vd 'Ngày làm việc/\\nWorking
    Date'), dong 2 = tieng Viet (nghieng).
    Neu chi co 1 trong 2 chuoi -> chi ghi chuoi do (van dung font thuong).
    Neu ca hai deu rong -> tra ve None (o se duoc de TRONG, xem clear_or_fill).

    QUAN TRONG: `size`/`font_name` PHAI duoc truyen vao tu font THAT cua o
    dich (vd template dung Times New Roman 20pt cho cot B, 24pt cho o Vi
    tri) — khong duoc de mac dinh Arial 11pt, neu khong chu se in ra qua
    NHO so voi phan con lai cua bao cao (loi da gap thuc te)."""
    en = _sanitize_xml_text((en or "").strip())
    vi = _sanitize_xml_text((vi or "").strip())
    if not en and not vi:
        return None
    fn = font_name or FONT_NAME
    normal = InlineFont(rFont=fn, sz=size, b=bold, i=False)
    italic = InlineFont(rFont=fn, sz=size, b=bold, i=True)
    parts = []
    if en:
        parts.append(TextBlock(normal, en + (" /" if vi else "")))
    if en and vi:
        parts.append("\n")
    if vi:
        parts.append(TextBlock(italic, vi))
    return CellRichText(*parts)


# --------------------------------------------------------------------------
# 2) DICH TU DONG KHI CHI NHAP 1 NGON NGU
# --------------------------------------------------------------------------

_VI_CHARS = re.compile(r"[ăâđêôơưÁÀẢÃẠăắằẳẵặÂấầẩẫậêếềểễệôốồổỗộơớờởỡợưứừửữựĐ]", re.IGNORECASE)


def detect_lang(text: str) -> str:
    """Doan ngon ngu don gian dua tren dau tieng Viet. Du du dung cho form
    nhap lieu cong truong (khong can thu vien NLP nang)."""
    if _VI_CHARS.search(text or ""):
        return "vi"
    return "en"


def build_glossary(*workbook_paths):
    """Hoc tu dien Anh<->Viet THAT tu chinh cac file mau da co (cac o dang
    'A / \\nB' hoac 'A/\\nB'). Day la nguon tu vung dang tin cay nhat vi no
    la thuat ngu du an that, khong phai tu dien chung chung."""
    glossary = {}
    pattern = re.compile(r"^(.*?)\s*/\s*\n?\s*(.*)$", re.DOTALL)
    for path in workbook_paths:
        try:
            wb = openpyxl.load_workbook(path, data_only=False)
        except Exception:
            continue
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not isinstance(v, str) or "/" not in v or "\n" not in v:
                        continue
                    m = pattern.match(v.strip())
                    if not m:
                        continue
                    a, b = m.group(1).strip(), m.group(2).strip()
                    if not a or not b or len(a) > 80 or len(b) > 80:
                        continue
                    # xac dinh ai la Anh, ai la Viet
                    if detect_lang(a) == "vi" and detect_lang(b) == "en":
                        vi_txt, en_txt = a, b
                    elif detect_lang(a) == "en" and detect_lang(b) == "vi":
                        en_txt, vi_txt = a, b
                    else:
                        continue
                    glossary[en_txt.lower()] = vi_txt
                    glossary[vi_txt.lower()] = en_txt
    return glossary


def external_translate(text: str, target_lang: str) -> Optional[str]:
    """*** DIEM NOI VOI DICH VU DICH THAT KHI TRIEN KHAI PRODUCTION ***
    Trong app that, ham nay se goi API dich (vi du Claude API / Google
    Translate) va tra ve chuoi da dich. O ban demo nay, moi truong sandbox
    khong co mang nen ham tra ve None (nghia la "khong dich duoc") va
    caller se roi xuong glossary/placeholder.
    """
    return None


def translate(text: str, glossary: dict) -> str:
    """Dich 1 chieu dung glossary hoc duoc; neu khong co trong tu dien,
    thu goi external_translate(); neu van khong duoc, tra ve placeholder
    de nguoi dung/bien tap vien biet la CAN DICH THEM (khong bao gio tu
    bia noi dung)."""
    text = (text or "").strip()
    if not text:
        return None
    src = detect_lang(text)
    tgt = "vi" if src == "en" else "en"
    hit = glossary.get(text.lower())
    if hit:
        return hit
    ext = external_translate(text, tgt)
    if ext:
        return ext
    return None  # KHONG dich duoc: tra None, KHONG bao gio bia chu "[cần dịch: ...]"
    # vao noi dung that cua o (loi da gap: chu bia bien thanh "dong thua" xuat
    # hien ngay tren file in that). Xem make_bilingual() ben duoi de biet
    # cach xu ly khi khong dich duoc: chi ghi 1 ngon ngu, danh dau bang
    # COMMENT rieng, khong dung noi dung o.


def make_bilingual(en_input: str, vi_input: str, glossary: dict, size=11, bold=False,
                    font_name: str = None):
    """Dau vao co the: (1) ca 2 ngon ngu da nhap -> dung nguyen,
    (2) chi 1 trong 2 -> THU tu dong dich ra ngon ngu con lai,
    (3) ca hai rong -> tra ve (None, False) (o se de trong).

    Tra ve tuple (rich_text_or_str, needs_review: bool). needs_review=True
    nghia la KHONG dich duoc (ngoai tuyen) -> o CHI chua 1 ngon ngu nguoi
    dung da go (khong bia them dong thu 2), va caller nen gan 1 CELL
    COMMENT de nguoi rieng biet nhung khong lam xau noi dung in ra.

    `size`/`font_name`: PHAI la font THAT cua o dich (xem ghi chu trong
    bilingual_richtext) — caller doc tu cell.font truoc khi goi ham nay."""
    en_input = _sanitize_xml_text((en_input or "").strip())
    vi_input = _sanitize_xml_text((vi_input or "").strip())
    if not en_input and not vi_input:
        return None, False

    needs_review = False
    if en_input and not vi_input:
        translated = translate(en_input, glossary)
        if translated:
            vi_input = _sanitize_xml_text(translated)
        else:
            needs_review = True
    elif vi_input and not en_input:
        translated = translate(vi_input, glossary)
        if translated:
            en_input = _sanitize_xml_text(translated)
        else:
            needs_review = True

    if needs_review:
        # chi 1 ngon ngu duoc xac nhan -> ghi DUNG 1 dong, khong bia dong 2
        only_text = en_input or vi_input
        only_is_vi = bool(vi_input) and not en_input
        font = InlineFont(rFont=font_name or FONT_NAME, sz=size, b=bold, i=only_is_vi)
        return CellRichText(TextBlock(font, only_text)), True

    return bilingual_richtext(en_input, vi_input, size=size, bold=bold, font_name=font_name), False


# --------------------------------------------------------------------------
# 3) + 4) O NHAP LIEU (to mau / chu do trong template) -> nen trang/chu den
#         khi xuat; o KHONG nhap thi de TRONG (khong lay mau demo cua template)
# --------------------------------------------------------------------------

INPUT_FILL_COLORS = {"FFFFFF00", "FFFFFF00".lower()}   # vang = o nhap tay
INPUT_FONT_COLORS = {"FFFF0000", "FFFF0000".lower()}   # chu do = da dien tay len nen tieu de

WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
BLACK_FONT_KW = dict(color="FF000000")

from openpyxl.styles.colors import COLOR_INDEX


def resolve_color(color) -> Optional[str]:
    """Quy toan bo kieu mau cua openpyxl (rgb / indexed / theme) ve 1 chuoi
    ARGB 8 ky tu de so sanh. QUAN TRONG: nhieu file Excel xuat tu ban cu
    (hoac luu qua nhieu lan) dung 'indexed color' (bang mau legacy 56 mau)
    thay vi rgb truc tiep -> neu chi doc .rgb se BO SOT o (vi du: o K2 cua
    file mau that dung font indexed=10, tuong duong FF0000, nhung .rgb tra
    ve None). Ham nay dam bao khong bo sot cac o nhu vay."""
    if color is None:
        return None
    ctype = getattr(color, "type", None)
    if ctype == "rgb" and isinstance(color.rgb, str):
        return color.rgb.upper()
    if ctype == "indexed":
        idx = color.indexed
        if idx is not None and 0 <= idx < len(COLOR_INDEX):
            return COLOR_INDEX[idx].upper()
    # theme color: khong the tra chinh xac gia tri that neu thieu theme.xml;
    # tra None (khong danh dau la input) thay vi doan sai.
    return None


def _is_input_style(cell) -> bool:
    fg = resolve_color(cell.fill.fgColor) if cell.fill else None
    fc = resolve_color(cell.font.color) if cell.font else None
    is_yellow = isinstance(fg, str) and fg.endswith("FFFF00")
    is_red = isinstance(fc, str) and fc.endswith("FF0000")
    return is_yellow or is_red


def _is_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def find_input_cells(ws):
    """Quet toan bo sheet, tra ve list cac cell duoc TEMPLATE danh dau la
    'vung nhap lieu' (to vang hoac chu do - ke ca khi mau duoc luu duoi
    dang indexed color), theo dung quy uoc da phat hien trong file mau
    that (khong hardcode toa do -> dung duoc cho moi sheet/moi file cung
    quy uoc). MOI o KHAC (khong nam trong danh sach nay) duoc xem la noi
    dung TINH / CO DINH cua template va se KHONG BAO GIO bi dong nay dong
    cham toi -> gia tri goc duoc giu nguyen 100%."""
    found = []
    for row in ws.iter_rows():
        for cell in row:
            if _is_input_style(cell):
                found.append(cell.coordinate)
    return found


def reset_input_style(cell):
    """Quy tac 3: xuat ra Excel -> o nhap lieu phai nen TRANG, chu DEN
    (khong con vang/do de 'nhac nho nhap lieu' nua vi du lieu da co)."""
    cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
    f = cell.font
    cell.font = Font(
        name=f.name or FONT_NAME,
        sz=f.sz or 11,
        b=f.b,
        i=f.i,
        color="FF000000",
    )


def apply_field(ws, coord: str, value, keep_formula_if_empty=True, needs_review=False):
    """Ghi 1 gia tri (str / number / CellRichText / None) vao o `coord`.
    - value is None/"" -> QUY TAC 4: neu o dang la cong thuc (tinh toan tu
      dong nhu SUM/HLOOKUP) thi GIU NGUYEN cong thuc (day khong phai o
      "nguoi dung nhap" ma la o "he thong tinh"); neu la o nhap tay thuan
      tuy -> xoa trang, KHONG lay lai gia tri mau cua template.
    - value co du lieu -> ghi de, roi luon reset mau nen/chu (Quy tac 3).
    - needs_review=True -> KHONG dich duoc offline: gan 1 CELL COMMENT
      (chu thich, chi hien khi ren chuot vao o trong Excel) thay vi bia
      chu vao noi dung — giu trang in SACH, khong co "dong thua"."""
    cell = ws[coord]
    if value is None or value == "":
        if keep_formula_if_empty and _is_formula(cell):
            pass  # giu cong thuc, khong dong gi them
        else:
            cell.value = None
    else:
        cell.value = value
    reset_input_style(cell)
    if needs_review:
        cell.comment = Comment(
            "⚠ Chỉ có 1 ngôn ngữ được xác nhận — cần bổ sung bản dịch còn lại "
            "(dịch tự động ngoại tuyến không khả dụng lúc xuất file này).",
            "report_engine.py",
        )
    return cell


def clear_all_inputs_first(ws, input_coords):
    """Buoc bat buoc TRUOC khi ghi du lieu that: xoa het gia tri MAU/DEMO
    co san trong file template (vd 'Shaft 42.15-42.18', 'Sunny'...), de
    dam bao khi APP xuat bao cao moi, khong o nao con sot du lieu cu."""
    for coord in input_coords:
        cell = ws[coord]
        if not _is_formula(cell):
            cell.value = None
        reset_input_style(cell)


# --------------------------------------------------------------------------
# 5) TRANG ANH: tu dong nhan ban khoi (1 dong mo ta + 2 khung anh) / 2 anh,
#    tu dong them trang khi vuot qua so khung co san
# --------------------------------------------------------------------------

# Hinh hoc phat hien duoc tu chinh file mau that (sheet "42.15-42.18"):
#   - Khoi lap lai moi 33 dong: 1 dong caption (merge B:C, cao ~45.75pt,
#     wrap text) + 2 khung anh, moi khung cao 16 dong x ~22.95pt.
#   - Khung anh chiem cot A->C (anchor tu giua cot A den giua cot C).
#   - print_area va ngat trang thu cong duoc dat sau MOI khoi (2 anh/trang in).
CAPTION_ROW_HEIGHT = 45.75
FRAME_ROW_HEIGHT = 22.95
FRAME_ROWS = 16
BLOCK_ROWS = 1 + FRAME_ROWS * 2       # 33
FIRST_BLOCK_CAPTION_ROW = 5           # dong caption dau tien trong template
FRAME_COL_FIRST = "A"
FRAME_COL_LAST = "C"
# So khoi anh CO SAN san trong template (dong 5, 38, 71 = 3 khoi, cho
# TOI DA 6 anh) — biet truoc de don dep du lieu mau con sot lai o cac
# khoi KHONG duoc dung toi trong lan xuat nay (vd o B38 co san so
# "12345" la du lieu demo cua nguoi lam template, phai xoa neu khong
# dung khoi do, neu khong no se "khong doi" moi lan xuat report moi).
PREBUILT_BLOCK_COUNT = 3


@dataclass
class Photo:
    path: str
    caption_en: str = ""
    caption_vi: str = ""


class PhotoPageBuilder:
    """Sinh (hoac tai su dung) cac khoi anh tren sheet anh, gan anh that,
    dien caption song ngu, va tu dong nhan them khoi/ngat trang khi so
    anh vuot qua suc chua hien co cua template."""

    def __init__(self, ws, glossary: dict):
        self.ws = ws
        self.glossary = glossary

    # -- utility: doc/ghi style toan bo 1 dong de nhan ban --------------
    def _copy_row_style(self, src_row: int, dst_row: int, n_cols=6):
        src_dim = self.ws.row_dimensions[src_row]
        dst_dim = self.ws.row_dimensions[dst_row]
        dst_dim.height = src_dim.height
        for c in range(1, n_cols + 1):
            s = self.ws.cell(row=src_row, column=c)
            d = self.ws.cell(row=dst_row, column=c)
            d.font = copy.copy(s.font)
            d.fill = copy.copy(s.fill)
            d.border = copy.copy(s.border)
            d.alignment = copy.copy(s.alignment)
            d.number_format = s.number_format

    def _ensure_block(self, block_index: int):
        """Bao dam khoi thu `block_index` (0-based) ton tai voi dung dinh
        dang (dong cao, merge caption...). Khoi 0/1/2 co san trong template;
        tu khoi thu 3 tro di se duoc NHAN BAN tu khoi mau (khoi 0)."""
        cap_row = FIRST_BLOCK_CAPTION_ROW + block_index * BLOCK_ROWS
        if self.ws.cell(row=cap_row, column=1).coordinate not in [
            m.coord for m in []
        ]:
            pass  # (giu placeholder ro rang cho phan mo rong sau nay)

        template_cap_row = FIRST_BLOCK_CAPTION_ROW
        already_has_rowdim = cap_row in self.ws.row_dimensions and \
            self.ws.row_dimensions[cap_row].height is not None

        if not already_has_rowdim:
            # chen moi 33 dong trang, sao chep dinh dang tu khoi mau (khoi 0)
            self.ws.insert_rows(cap_row, amount=BLOCK_ROWS)
            for offset in range(BLOCK_ROWS):
                self._copy_row_style(template_cap_row + offset, cap_row + offset)
            # merge caption B:C cua khoi moi
            self.ws.merge_cells(
                start_row=cap_row, start_column=2, end_row=cap_row, end_column=3
            )
            self.ws.row_dimensions[cap_row].height = CAPTION_ROW_HEIGHT
            self.ws.cell(row=cap_row, column=1).value = "Mô tả/ Description:"

            # dat ngat trang SAU khoi truoc do (moi khoi = 1 trang in = 2 anh)
            prev_last_row = cap_row - 1
            self.ws.row_breaks.append(Break(id=prev_last_row))

        return cap_row

    def _frame_anchor_rows(self, cap_row: int, slot: int):
        """slot 0 -> khung anh tren, slot 1 -> khung anh duoi, trong 1 khoi."""
        start = cap_row + 1 + slot * FRAME_ROWS
        end = start + FRAME_ROWS - 1
        return start, end

    def _two_cell_anchor(self, start_row_1idx: int, end_row_1idx: int):
        """QUAN TRONG — day la fix cho loi 'anh tran ra ngoai khung':
        truoc day anh duoc gan kich thuoc CO DINH theo PX uoc luong tu
        chieu cao dong (vd 16 dong x 22.95pt). Nhung neu file duoc mo lai
        boi LibreOffice/Excel va chieu cao dong bi lam tron nhe (da gap
        thuc te: 22.95pt -> 22.5pt sau khi recalc.py resave), anh CO KICH
        THUOC PX CO DINH se khong con khop dong nua -> tran sang khung anh
        ke ben.
        Cach fix DUNG: dung TwoCellAnchor gan thang vao toa do o (A{start}
        -> ngay-truoc-D{end+1}), KHONG dua tren px. Excel/LibreOffice se
        luon ve anh vua khit tu vien tren-trai den vien duoi-phai cua vung
        o do, bat ke chieu cao/rong dong-cot thuc te la bao nhieu — giong
        y cach chinh file mau goc tu lam (cung dung twoCellAnchor)."""
        from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker

        col_from = column_index_from_string(FRAME_COL_FIRST) - 1          # A -> 0
        col_to = column_index_from_string(FRAME_COL_LAST) - 1 + 1         # ngay sau C -> 3 (D)
        row_from = start_row_1idx - 1
        row_to = end_row_1idx                                             # ngay sau dong cuoi (0-idx)

        _from = AnchorMarker(col=col_from, colOff=0, row=row_from, rowOff=0)
        _to = AnchorMarker(col=col_to, colOff=0, row=row_to, rowOff=0)
        return TwoCellAnchor(editAs="oneCell", _from=_from, to=_to)

    def place_photo(self, photo: Photo, index: int):
        """index: 0-based thu tu anh trong toan bo bao cao."""
        block_index = index // 2
        slot = index % 2
        cap_row = self._ensure_block(block_index)

        if slot == 0:
            cap_cell = self.ws.cell(row=cap_row, column=2)
            existing_font = cap_cell.font
            cell_size = existing_font.sz or 11
            cell_font_name = existing_font.name or FONT_NAME
            rt, needs_review = make_bilingual(photo.caption_en, photo.caption_vi, self.glossary,
                                               size=cell_size, font_name=cell_font_name, bold=False)
            if rt is not None:
                cap_cell.value = rt
            if needs_review:
                cap_cell.comment = Comment(
                    "⚠ Chỉ có 1 ngôn ngữ được xác nhận cho mô tả ảnh này — "
                    "cần bổ sung bản dịch còn lại.", "report_engine.py")
            cap_cell.alignment = copy.copy(cap_cell.alignment)
            cap_cell.alignment = cap_cell.alignment.copy(wrap_text=True, vertical="center")
            autofit_caption_row(self.ws, cap_row, str(photo.caption_en) + str(photo.caption_vi))

        start_row, end_row = self._frame_anchor_rows(cap_row, slot)

        img = XLImage(photo.path)
        img.anchor = self._two_cell_anchor(start_row, end_row)
        self.ws.add_image(img)

    def finalize(self, n_photos: int):
        last_block = max(0, (n_photos - 1) // 2) if n_photos else -1
        last_row = FIRST_BLOCK_CAPTION_ROW + (last_block + 1) * BLOCK_ROWS - 1

        # *** FIX LOI: "dong 5/38/71 khong doi khi nhap thong tin" ***
        # Don sach caption + noi dung con sot cua CAC KHOI CO SAN nhung
        # KHONG duoc dung toi lan nay (vd chi co 2 anh -> chi dung khoi 0,
        # khoi 1(dong38)/2(dong71) phai duoc XOA TRANG, khong duoc de lai
        # du lieu demo cu cua template).
        for block_index in range(last_block + 1, PREBUILT_BLOCK_COUNT):
            cap_row = FIRST_BLOCK_CAPTION_ROW + block_index * BLOCK_ROWS
            if cap_row in self.ws.row_dimensions and self.ws.row_dimensions[cap_row].height is not None:
                cap_cell = self.ws.cell(row=cap_row, column=2)
                cap_cell.value = None
                if cap_cell.comment:
                    cap_cell.comment = None

        if last_row < FIRST_BLOCK_CAPTION_ROW:
            last_row = FIRST_BLOCK_CAPTION_ROW + BLOCK_ROWS - 1
        self.ws.print_area = f"A1:C{last_row}"
        # ngat trang cuoi cung
        self.ws.row_breaks.append(Break(id=last_row))


# --------------------------------------------------------------------------
# 6) TU DONG WRAP + CHINH CHIEU CAO DONG MO TA (khong de chu de len trang in)
# --------------------------------------------------------------------------

def autofit_caption_row(ws, row: int, sample_text: str, chars_per_line=70, line_height=15.0):
    """Uoc tinh so dong can thiet dua tren do dai text (ca EN+VI) va be
    rong cot B:C, tu do NANG chieu cao dong len neu can - tranh tinh trang
    chu tran ra ngoai khung khi in (Quy tac 6)."""
    text_len = len(sample_text or "")
    n_lines = max(1, -(-text_len // chars_per_line))  # ceil
    needed = max(CAPTION_ROW_HEIGHT, n_lines * line_height + 10)
    cur = ws.row_dimensions[row].height or CAPTION_ROW_HEIGHT
    if needed > cur:
        ws.row_dimensions[row].height = needed


# --------------------------------------------------------------------------
# QUY TRINH TONG
# --------------------------------------------------------------------------
#
# BO SUNG THEO YEU CAU: xuat ra 2 FILE EXCEL RIENG BIET, khong gop chung:
#   - generate_daily_report()  -> tu template Original1 (2 sheet: form +
#                                  trang anh). CO dinh kem trang anh.
#   - generate_daily_plan()    -> tu template Original2 (sheet ke hoach).
#                                  KHONG dinh kem anh (khong goi
#                                  PhotoPageBuilder, khong tao/giu drawing
#                                  nao trong file nay du photos co truyen
#                                  vao hay khong).
#
# CA HAI deu dung chung nguyen tac: chi cac o duoc TEMPLATE danh dau la
# "vung nhap lieu" (to vang / chu do - xem find_input_cells) moi bi dong
# nay dong toi (xoa-neu-trong, ghi-du-lieu, doi mau nen/chu). MOI o con
# lai (nhan/tieu de/khung bang... dang la chu den nen trang san co trong
# template) KHONG NAM trong danh sach do -> vong lap khong bao gio chay
# qua no -> gia tri va dinh dang GOC duoc giu nguyen 100%.
# --------------------------------------------------------------------------

# --------------------------------------------------------------------------
# QUY TRINH TONG
# --------------------------------------------------------------------------
#
# CAP NHAT QUAN TRONG (sau khi doi chieu DR1/DR2 that voi ban Empty cua
# chung): file DR1.xlsx / DR2.xlsx thuc te KHONG con dung quy uoc to vang/
# chu do nua (da kiem tra: fill_type=None, font mau theme/indexed thuong,
# khong phai vang/do). Vi vay find_input_cells() (do mau) KHONG con dang
# tin cho 2 file nay — ta chuyen sang dung 1 DANH SACH TOA DO TUONG MINH,
# da duoc kiem chung bang cach so sanh truc tiep ban co du lieu voi ban
# trong (xem json_to_excel.py: REPORT_FIELDS / PLAN_FIELDS).
#
# Danh sach tuong minh nay con phan biet 2 loai o:
#   - "reset" o: LUON xoa trang truoc (neu app khong gui gi thi o do PHAI
#     trong that trong bao cao moi) — vd so luong, gio bat dau/ket thuc,
#     ngay thang... (Quy tac 4)
#   - "preserve" o: CHI ghi de KHI co du lieu moi; neu app khong gui gi,
#     GIU NGUYEN gia tri san co cua template (vd ten hang muc cong viec
#     "Normal Pipe Jacking Works..." la cau hinh co dinh cua du an, khong
#     phai o rong-cho-nhap — khong duoc xoa no chi vi app khong gui lai
#     moi ngay).
#
# Neu KHONG truyen input_map (dung cho template khac chua kiem chung),
# ham roi ve hanh vi cu: dung find_input_cells() (do mau) nhu truoc.
# --------------------------------------------------------------------------

def autofit_row_for_cell(ws, coord: str, text_len: int, font_size: float = 11):
    """Ban tong quat cua autofit_caption_row — dung cho BAT KY o nao (khong
    chi caption anh): uoc luong so dong can dua tren do rong cot THAT va
    co chu THAT, tu dong NANG chieu cao dong len neu can, tranh chu bi
    tran/de len dong ke tiep khi in (Quy tac 6 - khong de 'chen chu')."""
    from openpyxl.utils.cell import coordinate_from_string
    col_letter, row = coordinate_from_string(coord)
    dim = ws.column_dimensions.get(col_letter)
    col_width = dim.width if (dim and dim.width) else 10
    chars_per_line = max(int(col_width * 1.8), 5)
    n_lines = max(1, -(-text_len // chars_per_line))
    line_height = max(font_size * 1.3, 15.0)
    needed = n_lines * line_height + 6
    cur = ws.row_dimensions[row].height or 15.0
    if needed > cur:
        ws.row_dimensions[row].height = needed


def _wrap_text_cell(cell):
    cell.alignment = cell.alignment.copy(wrap_text=True)


def _apply_data_sheets(wb, data: dict, glossary: dict, input_map: dict = None):
    """Ghi du lieu vao cac sheet du lieu (dung chung cho ca Daily Report
    lan Daily Plan).

    input_map (neu co): {sheet_name: {"reset": [...toa do...], "preserve": [...toa do...]}}
    """
    touched = {}
    for sheet_name, fields in data.items():
        ws = wb[sheet_name]

        if input_map and sheet_name in input_map:
            reset_coords = set(input_map[sheet_name].get("reset", []))
            preserve_coords = set(input_map[sheet_name].get("preserve", []))
        else:
            reset_coords = set(find_input_cells(ws))
            preserve_coords = set()
        allowed = reset_coords | preserve_coords
        clear_all_inputs_first(ws, reset_coords)   # CHI xoa nhom "reset"

        for coord, val in fields.items():
            if coord not in allowed:
                # An toan: tu choi ghi ra ngoai vung da duoc kiem chung la
                # o nhap lieu, de khong bao gio vo tinh de len noi dung tinh.
                continue
            is_empty = val is None or val == "" or (isinstance(val, dict) and not (val.get("en") or "").strip() and not (val.get("vi") or "").strip())
            if coord in preserve_coords and is_empty:
                continue  # "preserve": khong gui gi -> giu nguyen, khong dong cham
            if isinstance(val, dict):
                # QUAN TRONG: doc font THAT cua o dich TRUOC khi ghi de -
                # template co the dung Times New Roman 20pt/24pt (da xac
                # nhan thuc te), khong phai Arial 11pt mac dinh. Neu
                # khong lam buoc nay, chu se in ra qua nho (loi da gap).
                existing_font = ws[coord].font
                cell_size = existing_font.sz or 11
                cell_font_name = existing_font.name or FONT_NAME
                rt, needs_review = make_bilingual(
                    val.get("en"), val.get("vi"), glossary,
                    size=cell_size, font_name=cell_font_name,
                )
                if rt is None:
                    continue
                cell = apply_field(ws, coord, rt, needs_review=needs_review)
                _wrap_text_cell(cell)
                combined_len = len((val.get("en") or "")) + len((val.get("vi") or ""))
                autofit_row_for_cell(ws, coord, combined_len, font_size=cell_size)
            else:
                apply_field(ws, coord, val)
        touched[sheet_name] = allowed
    return touched


def strip_legacy_photo_shapes(xlsx_path: str):
    """*** FIX LOI: 'trang anh khong doi khi nhap thong tin' ***
    Nguyen nhan that: cac khung anh MAU trong template (dong 5/38/71) KHONG
    phai anh that (<xdr:pic>) ma la SHAPE hinh chu nhat duoc to day bang
    anh qua blipFill (<xdr:sp> chua <a:blipFill>) — day la cach nguoi lam
    template cu chen "anh minh hoa" ma khong dung tinh nang Insert Picture
    chuan. Khi minh chen anh MOI (dung dung <xdr:pic>), shape gia nay VAN
    CON NGUYEN trong file, nam chong len dung vi tri anh moi -> nhin như
    "khong doi" (shape cu de len tren, hoac ca 2 lop cung hien).

    Ham nay mo lai file .xlsx SAU KHI luu, tim tat ca <xdr:sp> co chua
    <a:blipFill> ben trong (dung la dau hieu "anh gia dang shape"), va
    xoa CHINH XAC nhung shape do — KHONG dong gi den cac shape trang tri
    khac (o hop van ban rong...) hay anh that (<xdr:pic>) minh vua chen."""
    NS = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    }
    drawing_pattern = re.compile(r"^xl/drawings/drawing\d+\.xml$")
    tmp_path = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        drawing_names = [n for n in zin.namelist() if drawing_pattern.match(n)]
        if not drawing_names:
            return
        replacements = {}
        for dn in drawing_names:
            content = zin.read(dn)
            try:
                root = ET.fromstring(content)
            except ET.ParseError:
                continue
            changed = False
            for anchor in list(root):
                sp = anchor.find("xdr:sp", NS)
                if sp is not None and sp.find(".//a:blipFill", NS) is not None:
                    root.remove(anchor)
                    changed = True
            if changed:
                replacements[dn] = ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
        if not replacements:
            return
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = replacements.get(item.filename, zin.read(item.filename))
                zout.writestr(item, data)
    os.replace(tmp_path, xlsx_path)


def generate_daily_report(template_path: str, output_path: str, data: dict,
                           photos: list, photo_sheet_name: str = "42.15-42.18",
                           input_map: dict = None):
    """Xuat FILE 1: Daily Report — co dinh kem trang anh.
    `data` chi chua sheet(s) du lieu (KHONG bao gom sheet anh)."""
    glossary = build_glossary(template_path)
    wb = openpyxl.load_workbook(template_path)

    _apply_data_sheets(wb, data, glossary, input_map=input_map)

    # LUON xu ly sheet anh (ke ca khi photos=[]) de dam bao du lieu demo
    # con sot cua template (vd B38="12345") luon duoc don sach, khong chi
    # khi co anh moi.
    if photo_sheet_name not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy sheet ảnh '{photo_sheet_name}' trong template.")
    ws_photo = wb[photo_sheet_name]
    builder = PhotoPageBuilder(ws_photo, glossary)
    for i, ph in enumerate(photos):
        builder.place_photo(ph, i)
    builder.finalize(len(photos))

    wb.save(output_path)
    strip_legacy_photo_shapes(output_path)   # xoa anh-gia-dang-shape con sot
    return output_path


def generate_daily_plan(template_path: str, output_path: str, data: dict, input_map: dict = None):
    """Xuat FILE 2: Daily Plan — KHONG dinh kem anh. Ham nay khong nhan
    tham so `photos` va khong bao gio goi PhotoPageBuilder, de dam bao
    file ke hoach luon la file thuan du lieu, tach biet hoan toan voi
    file Daily Report co anh."""
    glossary = build_glossary(template_path)
    wb = openpyxl.load_workbook(template_path)
    _apply_data_sheets(wb, data, glossary, input_map=input_map)
    wb.save(output_path)
    return output_path


def assert_static_cells_preserved(original_path: str, generated_path: str,
                                   sheet_name: str, input_coords: set):
    """Ham kiem chung (dung trong test / CI): doi chieu TOAN BO cac o
    KHONG nam trong `input_coords` giua file goc va file vua xuat — neu
    bat ky o tinh nao bi doi gia tri, ham se raise loi ngay. Dung de
    chung minh yeu cau 'phan chu den nen trang con lai giu nguyen so
    lieu' duoc tuan thu, khong chi 'noi suong'."""
    wb_o = openpyxl.load_workbook(original_path)
    wb_g = openpyxl.load_workbook(generated_path)
    ws_o, ws_g = wb_o[sheet_name], wb_g[sheet_name]
    mismatches = []
    for row in ws_o.iter_rows():
        for cell in row:
            if cell.coordinate in input_coords:
                continue
            gv = ws_g[cell.coordinate].value
            if cell.value != gv:
                mismatches.append((cell.coordinate, cell.value, gv))
    if mismatches:
        raise AssertionError(
            f"{len(mismatches)} ô TĨNH bị thay đổi ngoài ý muốn: {mismatches[:10]}"
        )
    return True
