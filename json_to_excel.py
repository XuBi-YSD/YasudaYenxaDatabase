"""
json_to_excel.py
=================
Cau noi giua APP (xuat file .json) va report_engine.py (sinh file .xlsx
that). Ban nay duoc VIET LAI HOAN TOAN sau khi doi chieu truc tiep 2 cap
file that (20260731-DR1.xlsx / DR2.xlsx) voi ban trong tuong ung
(..._Empty.xlsx) — moi toa do o duoi day deu duoc XAC MINH bang bang
chung thuc te (khong con doan theo mau sac, vi DR1/DR2 KHONG con dung
quy uoc to vang/chu do nua).

CHAY:
    python json_to_excel.py report --template 20260731-DR1.xlsx \
        --payload daily_report_payload.json --output out.xlsx

    python json_to_excel.py plan --template 20260731-DR2.xlsx \
        --payload daily_plan_payload.json --output out.xlsx
"""

import argparse
import base64
import datetime as dt
import json
import re
import sys
import tempfile
from pathlib import Path

from report_engine import generate_daily_report, generate_daily_plan, Photo

# ============================================================================
# FIELD MAP — XAC MINH BANG CACH SO SANH FILE CO DU LIEU vs FILE TRONG
# ============================================================================
#
# Cau truc chung cho 1 "activity row" (dong cong viec hien truong):
#   unit          — don vi tinh (vd "Pipe", "LS")
#   design_qty    — KLTK / khoi luong thiet ke
#   qty_today     — (chi DR2/Plan) KLTH hom nay — 1 vai dong la so, 1 vai la
#                   cong thuc (VD E11 la '=(E9)/D9') -> engine tu bao ve cong thuc
#   manpower_qty  — so luong nhan luc theo loai (loai da co san, chi nhap so)
#   equip_qty     — so luong thiet bi theo loai (loai da co san, chi nhap so)
#   start / finish— gio bat dau / ket thuc
#   description   — CHI dung khi muon THAY hang muc co san hoac THEM hang
#                   muc moi o cac dong con trong; neu bo trong -> GIU
#                   NGUYEN mo ta san co cua template (loai "preserve").

REPORT_FIELDS = {
    "sheet": "Daily report for PJ",
    "date_coord": "K2",
    "location_coord": "N2",

    # Bang 1: CONG VIEC HIEN TRUONG — B8:F17 la 1 khoi thong nhat (da xac
    # minh qua so sanh Empty/Filled: E,F KHONG phai cong thuc co dinh cua
    # template — trong ban Empty ca E9:F17 va ca H14/N31... deu la None,
    # nghia la "cong thuc" nguoi lam demo go vao chi la VI DU su dung,
    # KHONG phai cau truc bat buoc -> coi la o nhap lieu binh thuong).
    "activity_rows": [8, 9, 10, 11, 12, 13, 14, 15, 16, 17],
    "preset_description_rows": [8, 9, 10, 11],   # B8 + 3 dong dau: description = "preserve"
    "activity_cols": {
        "description": "B", "unit": "C", "design_qty": "D",
        "qty_today": "E", "accumulated": "F", "remark": "M",
    },

    # Bang 2: NHAN LUC & THIET BI & GIO LAM VIEC — danh sach DOC LAP,
    # cung hang voi bang 1 nhung KHONG lien quan noi dung. Cot G/I la
    # NHAN co san (da xac minh: cac o co san gia tri thi CHI DOC; cac o
    # dang trong — vd G15:G17 — duoc PHEP nhap moi khi can bo sung loai
    # nhan luc/thiet bi moi).
    "manpower_equip_rows": [8, 9, 10, 11, 12, 13, 14, 15, 16, 17],
    "manpower_equip_cols": {
        "manpower_legend": "G", "manpower_qty": "H",
        "equip_legend": "I", "equip_qty": "J",
        "start": "K", "finish": "L",
    },

    "inspection_rows": [19, 20],
    "inspection_col": "B",
    "next_day_rows": [22, 23, 24, 25, 26, 27],
    "next_day_col": "J",
    "weather": {
        "morning_cond": "I31", "afternoon_cond": "J31", "evening_cond": "K31",  # N31 khong phai cong thuc co dinh, nhung khong dung trong app nay
        "temp_morning": "L31", "temp_afternoon": "M31",
        "rainfall": "O30", "water_level": "P30", "comment": "Q30",
    },
}

PLAN_FIELDS = {
    "sheet": "Daily Plan for PJ",
    "date_coord": "N2",
    "location_coord": "Q2",

    "activity_rows": [8, 9, 10, 11, 12, 13, 14, 15],
    "preset_description_rows": [8, 9, 10, 11],
    # LUU Y: DR2 cot F la O MERGE VOI E (E6:F6), khong phai cot du lieu
    # rieng -> dung G (KLLK/Accumulative) thay vi F.
    "activity_cols": {
        "description": "B", "unit": "C", "design_qty": "D",
        "qty_today": "E", "accumulated": "G", "remark": "P",
    },

    "manpower_equip_rows": [8, 9, 10, 11, 12, 13, 14, 15],
    "manpower_equip_cols": {
        "manpower_legend": "H", "manpower_qty": "I",
        "equip_legend": "K", "equip_qty": "L",
        "start": "N", "finish": "O",
    },

    "inspection_rows": [19, 20],
    "inspection_col": "B",
    "next_day_rows": [22, 23, 24, 25, 26, 27],
    "next_day_col": "K",   # DR2 dung cot K, khong phai J nhu DR1
    "weather": {
        "morning_cond": "K29", "afternoon_cond": "L29", "evening_cond": "N29",  # Q29 la cong thuc
        "temp_morning": "O29", "temp_afternoon": "P29",
        "rainfall": "R28", "water_level": "S28", "comment": "T28",
    },
}


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, (dt.date, dt.datetime)):
        return value
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", str(value))
    if not m:
        raise ValueError(f"Ngày không đúng định dạng YYYY-MM-DD: {value!r}")
    y, mo, d = map(int, m.groups())
    return dt.date(y, mo, d)


def _parse_time(value):
    """'HH:MM' (tu <input type=time>) -> datetime.time THAT (khong de string,
    tranh loi so sanh/tinh toan gio trong Excel)."""
    if not value:
        return None
    if isinstance(value, dt.time):
        return value
    m = re.match(r"^(\d{1,2}):(\d{2})$", str(value))
    if not m:
        raise ValueError(f"Giờ không đúng định dạng HH:MM: {value!r}")
    h, mi = map(int, m.groups())
    return dt.time(h, mi)


def _bilingual_or_none(v):
    if not isinstance(v, dict):
        return v
    en, vi = (v.get("en") or "").strip(), (v.get("vi") or "").strip()
    if not en and not vi:
        return None
    return v


def _build_input_map(field_map: dict, template_path: str = None):
    """Sinh {"reset": [...], "preserve": [...]} tu FIELD_MAP — dung de
    truyen vao report_engine.generate_daily_report/plan(input_map=...),
    thay the hoan toan cho viec do mau (khong con dung duoc voi DR1/DR2).

    Neu co template_path: tu doc truc tiep cot nhan (G/I hoac H/K) de xac
    dinh o nao dang TRONG trong file mau that -> cho phep nhap moi (vd
    G15:G17 khi chua co san loai nhan luc) — o nao DA CO san nhan van
    CHI DOC nhu truoc."""
    reset, preserve = [], []
    reset += [field_map["date_coord"], field_map["location_coord"]]

    cols = field_map["activity_cols"]
    for row in field_map["activity_rows"]:
        for key, col in cols.items():
            coord = f"{col}{row}"
            if key == "description":
                if row in field_map["preset_description_rows"]:
                    preserve.append(coord)
                else:
                    reset.append(coord)
            else:
                reset.append(coord)

    me_cols = field_map["manpower_equip_cols"]
    editable_legend_rows = {"manpower_legend": set(), "equip_legend": set()}
    if template_path:
        for row in read_legend(template_path, field_map):
            if not row["manpower_legend"]:
                editable_legend_rows["manpower_legend"].add(row["row"])
            if not row["equip_legend"]:
                editable_legend_rows["equip_legend"].add(row["row"])

    for row in field_map["manpower_equip_rows"]:
        for key in ("manpower_qty", "equip_qty", "start", "finish"):
            reset.append(f"{me_cols[key]}{row}")
        # cot nhan (G/I hoac H/K): CHI cho nhap khi o do dang TRONG trong
        # template that (xac dinh dong luc o tren); neu da co san nhan
        # (vd G8="M") thi KHONG dua vao danh sach cho phep -> chi doc.
        if row in editable_legend_rows["manpower_legend"]:
            reset.append(f"{me_cols['manpower_legend']}{row}")
        if row in editable_legend_rows["equip_legend"]:
            reset.append(f"{me_cols['equip_legend']}{row}")

    for row in field_map["inspection_rows"]:
        reset.append(f"{field_map['inspection_col']}{row}")
    for row in field_map["next_day_rows"]:
        reset.append(f"{field_map['next_day_col']}{row}")
    for coord in field_map["weather"].values():
        reset.append(coord)

    return {"reset": reset, "preserve": preserve}


def build_data_dict(payload: dict, field_map: dict, template_path: str = None):
    """payload (tu app) -> (fields_by_coord, input_map, applied, skipped)."""
    sheet_name = field_map["sheet"]
    sp = payload.get("sheet_data", {}).get(sheet_name, {})
    fields = {}
    applied, skipped = [], []

    def _set(coord, val, label):
        if val is None or val == "":
            return
        fields[coord] = val
        applied.append(f"{label} -> {coord}")

    _set(field_map["date_coord"], _parse_date(sp.get("date")), "date")
    _set(field_map["location_coord"], _bilingual_or_none(sp.get("location")), "location")

    cols = field_map["activity_cols"]
    activities = sp.get("activities") or []
    rows = field_map["activity_rows"]
    if len(activities) > len(rows):
        raise ValueError(
            f"Có {len(activities)} dòng công việc nhưng template chỉ hỗ trợ "
            f"{len(rows)} dòng ({sheet_name}). Chèn thêm dòng cho bảng này "
            f"chưa được hỗ trợ (xem GIỚI HẠN ĐÃ BIẾT)."
        )
    for row_no, item in zip(rows, activities):
        for key, col in cols.items():
            raw = item.get(key)
            # Quy tac song ngu ap dung cho MOI truong (khong chi description):
            # neu gia tri la {'en':..,'vi':..} -> song ngu/tu dong dich;
            # neu la so/chuoi thuong -> giu nguyen (vd don vi "Pipe", so luong).
            val = _bilingual_or_none(raw) if isinstance(raw, dict) else raw
            if val in (None, ""):
                continue
            fields[f"{col}{row_no}"] = val
            applied.append(f"activity[row {row_no}].{key} -> {col}{row_no}")

    me_cols = field_map["manpower_equip_cols"]
    me_items = sp.get("manpower_equip") or []
    me_rows = field_map["manpower_equip_rows"]
    if len(me_items) > len(me_rows):
        raise ValueError(
            f"Có {len(me_items)} dòng nhân lực/thiết bị nhưng template chỉ hỗ trợ "
            f"{len(me_rows)} dòng ({sheet_name})."
        )
    for row_no, item in zip(me_rows, me_items):
        for key in ("manpower_qty", "equip_qty", "start", "finish", "manpower_legend", "equip_legend"):
            raw = item.get(key)
            if raw in (None, ""):
                continue
            val = _parse_time(raw) if key in ("start", "finish") else raw
            val = _bilingual_or_none(val) if isinstance(val, dict) else val
            if val in (None, ""):
                continue
            if key == "manpower_qty":
                # Dong dau tien (row8) = nhan luc Nhat Ban -> "(JP)";
                # cac dong con lai = nhan luc Viet Nam -> "(VN)" khi co so lieu.
                suffix = " (JP)" if row_no == me_rows[0] else " (VN)"
                val = f"{val}{suffix}"
            col = me_cols[key]
            fields[f"{col}{row_no}"] = val
            applied.append(f"manpower_equip[row {row_no}].{key} -> {col}{row_no}")

    insp = sp.get("inspections") or []
    for row_no, text in zip(field_map["inspection_rows"], insp):
        v = _bilingual_or_none(text)
        if v is not None:
            fields[f"{field_map['inspection_col']}{row_no}"] = v
            applied.append(f"inspection[row {row_no}] -> {field_map['inspection_col']}{row_no}")

    nextday = sp.get("next_day_work") or []
    for row_no, text in zip(field_map["next_day_rows"], nextday):
        v = _bilingual_or_none(text)
        if v is not None:
            fields[f"{field_map['next_day_col']}{row_no}"] = v
            applied.append(f"next_day_work[row {row_no}] -> {field_map['next_day_col']}{row_no}")

    w = sp.get("weather") or {}
    wm = field_map["weather"]
    for key in ("morning_cond", "afternoon_cond", "evening_cond"):
        _set(wm[key], w.get(key), f"weather.{key}")
    for key in ("temp_morning", "temp_afternoon", "rainfall", "water_level"):
        _set(wm[key], w.get(key), f"weather.{key}")
    _set(wm["comment"], _bilingual_or_none(w.get("comment")), "weather.comment")

    input_map = {sheet_name: _build_input_map(field_map, template_path=template_path)}
    return fields, input_map, applied, skipped


def _decode_photos(payload: dict, tmpdir: str):
    photos = []
    for i, p in enumerate(payload.get("photos", [])):
        data_url = p.get("dataUrl") or p.get("data_url")
        if not data_url:
            raise ValueError(f"Ảnh #{i} không có dữ liệu ảnh nhúng (dataUrl) trong JSON.")
        header, b64data = data_url.split(",", 1)
        ext = "jpg" if "jpeg" in header or "jpg" in header else "png"
        img_path = Path(tmpdir) / f"photo_{i}.{ext}"
        img_path.write_bytes(base64.b64decode(b64data))
        cap = p.get("caption") or {}
        photos.append(Photo(str(img_path), caption_en=cap.get("en", ""), caption_vi=cap.get("vi", "")))
    return photos


def read_legend(template_path: str, field_map: dict):
    """Doc CHINH XAC noi dung cot 'nhan' (G/I o DR1, H/K o DR2) tu file mau
    that — de app hien thi dung nhan tham chieu ben canh o nhap so luong,
    khong doan/hardcode trong Javascript (tranh lech voi file that)."""
    import openpyxl as _oxl
    wb = _oxl.load_workbook(template_path, data_only=True)
    ws = wb[field_map["sheet"]]
    cols = field_map["manpower_equip_cols"]
    out = []
    for row in field_map["manpower_equip_rows"]:
        out.append({
            "row": row,
            "manpower_legend": ws[f"{cols['manpower_legend']}{row}"].value,
            "equip_legend": ws[f"{cols['equip_legend']}{row}"].value,
        })
    return out


def convert_report(template, payload_path, output_path):
    payload = json.loads(Path(payload_path).read_text(encoding="utf-8"))
    fields, input_map, applied, skipped = build_data_dict(payload, REPORT_FIELDS, template_path=template)
    data = {REPORT_FIELDS["sheet"]: fields}
    with tempfile.TemporaryDirectory() as tmpdir:
        photos = _decode_photos(payload, tmpdir)
        generate_daily_report(template, output_path, data, photos, input_map=input_map)
    _print_summary("DAILY REPORT", output_path, applied, skipped, n_photos=len(photos))


def convert_plan(template, payload_path, output_path):
    payload = json.loads(Path(payload_path).read_text(encoding="utf-8"))
    fields, input_map, applied, skipped = build_data_dict(payload, PLAN_FIELDS, template_path=template)
    data = {PLAN_FIELDS["sheet"]: fields}
    if payload.get("photos"):
        raise ValueError(
            "Payload Daily Plan có chứa 'photos' nhưng file Daily Plan KHÔNG "
            "được đính kèm ảnh theo yêu cầu."
        )
    generate_daily_plan(template, output_path, data, input_map=input_map)
    _print_summary("DAILY PLAN", output_path, applied, skipped, n_photos=0)


def _print_summary(label, output_path, applied, skipped, n_photos):
    print(f"\n=== {label}: đã xuất → {output_path} ===")
    print(f"Ảnh đã chèn: {n_photos}")
    print(f"Trường đã ghi vào Excel ({len(applied)}):")
    for a in applied:
        print("  ✓", a)
    if skipped:
        print(f"Trường KHÔNG ghi được ({len(skipped)}):")
        for s in skipped:
            print("  ⚠", s)


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("kind", choices=["report", "plan"])
    ap.add_argument("--template", required=True)
    ap.add_argument("--payload", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    try:
        if args.kind == "report":
            convert_report(args.template, args.payload, args.output)
        else:
            convert_plan(args.template, args.payload, args.output)
    except Exception as e:
        print(f"\n LỖI: {e}", file=sys.stderr)
        sys.exit(1)
